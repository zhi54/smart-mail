#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel 读取模块

使用 openpyxl 或 xlrd 读取 Excel 工资条数据
"""

import os
from datetime import date
from openpyxl import load_workbook
import xlrd
from utils.logger import logger


class ExcelReader:
    """Excel 读取器"""

    # 字段映射：Excel列名 -> 数据字段
    FIELD_MAPPING = {
        '姓名': 'name',
        '应出勤天数': 'expected_days',
        '实际出勤天数': 'actual_days',
        '基本工资': 'base_salary',
        '绩效工资': 'performance_salary',
        '直播工资': 'live_salary',
        '提成': 'commission',
        '售后客服奖金': 'service_bonus',
        '税前工资': 'pre_tax_salary',
        '邮箱': 'email',
        '社保个人部分扣款': 'social_security',
        '住房公积金个人部分扣款': 'housing_fund',
        '专项抵扣': 'special_deduction',
        '本月扣除项累计': 'total_deduction',
        '累计应缴预缴所得额': 'accumulated_taxable',
        '累计税额': 'accumulated_tax',
        '本月应扣缴额': 'current_tax',
        '员工实得': 'net_salary',
        '发放月份': 'pay_month',
    }

    def __init__(self, file_path):
        """初始化 Excel 读取器

        Args:
            file_path: Excel 文件路径
        """
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
        self.headers = []
        self.data = []
        self.file_type = None
        self.has_pay_month = False  # 是否有发放月份列
        self.default_pay_month = self._get_default_pay_month()
        self._load()

    def _get_default_pay_month(self):
        """获取默认发放月份（上个月）

        Returns:
            格式化的月份字符串，如 "2025年12月"
        """
        today = date.today()
        # 计算上个月
        if today.month == 1:
            last_month = 12
            last_year = today.year - 1
        else:
            last_month = today.month - 1
            last_year = today.year
        return f"{last_year}年{last_month:02d}月"

    def _load(self):
        """加载 Excel 文件"""
        try:
            logger.info(f"正在加载 Excel 文件: {self.file_path}")

            # 根据文件扩展名选择读取方式
            if self.file_path.endswith('.xlsx'):
                self._load_xlsx()
            elif self.file_path.endswith('.xls'):
                self._load_xls()
            else:
                raise ValueError(f"不支持的文件格式: {self.file_path}")

            logger.info(f"成功读取 {len(self.data)} 条员工数据")

        except Exception as e:
            logger.error(f"加载 Excel 文件失败: {e}")
            raise

    def _load_xlsx(self):
        """加载 .xlsx 文件（使用 openpyxl）"""
        self.file_type = 'xlsx'
        self.workbook = load_workbook(self.file_path, data_only=True)
        self.sheet = self.workbook.active

        # 读取表头
        self.headers = [cell.value for cell in self.sheet[1]]
        logger.info(f"表头: {self.headers}")

        # 检查是否有发放月份列
        self.has_pay_month = '发放月份' in self.headers
        if not self.has_pay_month:
            logger.info(f"Excel 中没有'发放月份'列，将自动添加默认值: {self.default_pay_month}")

        # 读取数据
        self.data = self._parse_data_xlsx()

    def _load_xls(self):
        """加载 .xls 文件（使用 xlrd）"""
        self.file_type = 'xls'
        self.workbook = xlrd.open_workbook(self.file_path)
        self.sheet = self.workbook.sheet_by_index(0)

        # 读取表头
        self.headers = [self.sheet.cell_value(0, col) for col in range(self.sheet.ncols)]
        logger.info(f"表头: {self.headers}")

        # 检查是否有发放月份列
        self.has_pay_month = '发放月份' in self.headers
        if not self.has_pay_month:
            logger.info(f"Excel 中没有'发放月份'列，将自动添加默认值: {self.default_pay_month}")

        # 读取数据
        self.data = self._parse_data_xls()

    def _parse_data_xlsx(self):
        """解析 xlsx 数据"""
        data_list = []

        for row_idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not any(row):
                continue

            employee_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(self.headers):
                    header = self.headers[col_idx]
                    field_name = self.FIELD_MAPPING.get(header, header)

                    # 处理数字格式
                    if isinstance(value, (int, float)):
                        employee_data[field_name] = round(value, 2) if value else 0
                    else:
                        employee_data[field_name] = value if value else ''

            # 如果没有发放月份列，自动添加默认值
            if not self.has_pay_month:
                employee_data['pay_month'] = self.default_pay_month
            # 如果有发放月份列为空，也使用默认值
            elif not employee_data.get('pay_month'):
                employee_data['pay_month'] = self.default_pay_month

            # 验证必填字段
            if not employee_data.get('name') or not employee_data.get('email'):
                logger.warning(f"第 {row_idx} 行数据不完整，跳过")
                continue

            data_list.append(employee_data)

        return data_list

    def _parse_data_xls(self):
        """解析 xls 数据"""
        data_list = []

        for row_idx in range(1, self.sheet.nrows):
            row_values = self.sheet.row_values(row_idx)

            # 跳过空行
            if not any(row_values):
                continue

            employee_data = {}
            for col_idx, value in enumerate(row_values):
                if col_idx < len(self.headers):
                    header = self.headers[col_idx]
                    field_name = self.FIELD_MAPPING.get(header, header)

                    # 处理数字格式
                    if isinstance(value, (int, float)):
                        employee_data[field_name] = round(value, 2) if value else 0
                    else:
                        employee_data[field_name] = value if value else ''

            # 如果没有发放月份列，自动添加默认值
            if not self.has_pay_month:
                employee_data['pay_month'] = self.default_pay_month
            # 如果有发放月份列为空，也使用默认值
            elif not employee_data.get('pay_month'):
                employee_data['pay_month'] = self.default_pay_month

            # 验证必填字段
            if not employee_data.get('name') or not employee_data.get('email'):
                logger.warning(f"第 {row_idx + 1} 行数据不完整，跳过")
                continue

            data_list.append(employee_data)

        return data_list

    def get_data(self):
        """获取所有数据

        Returns:
            员工数据列表
        """
        return self.data

    def get_preview_data(self, count=3):
        """获取预览数据

        Args:
            count: 预览数量

        Returns:
            前 N 条员工数据
        """
        return self.data[:count]

    def get_total_count(self):
        """获取数据总数

        Returns:
            员工总数
        """
        return len(self.data)

    def get_headers(self):
        """获取表头

        Returns:
            表头列表
        """
        return self.headers

    def __len__(self):
        """获取数据数量"""
        return len(self.data)

    def __iter__(self):
        """迭代器"""
        return iter(self.data)
